import os
import pandas as pd

def find_value_in_excel(file_path, search_value="120/100"):
    """
    Ищет значение в Excel файле (поддерживает .xls и .xlsx)
    """
    if not os.path.exists(file_path):
        print(f"Файл не найден: {file_path}")
        return
    
    try:
        # Определяем формат по расширению
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xls':
            # Для старых .xls файлов используем xlrd
            import xlrd
            book = xlrd.open_workbook(file_path)
            
            print(f"Файл: {file_path}")
            print(f"Количество листов: {book.nsheets}")
            print(f"Листы: {book.sheet_names()}")
            
            found_count = 0
            search_str = str(search_value)
            
            for sheet_idx in range(book.nsheets):
                sheet = book.sheet_by_index(sheet_idx)
                print(f"\n--- Лист '{sheet.name}' ---")
                
                for rx in range(sheet.nrows):
                    for cx in range(sheet.ncols):
                        cell = sheet.cell(rx, cx)
                        if cell.ctype != xlrd.XL_CELL_EMPTY:
                            cell_value = str(cell.value)
                            if search_str in cell_value:
                                found_count += 1
                                print(f"  Строка {rx+1}, Столбец {cx+1}: {cell_value}")
            
            print(f"\nВсего найдено: {found_count}")
            
        elif file_ext == '.xlsx':
            # Для новых .xlsx файлов используем openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            
            print(f"Файл: {file_path}")
            print(f"Количество листов: {len(wb.sheetnames)}")
            print(f"Листы: {wb.sheetnames}")
            
            found_count = 0
            search_str = str(search_value)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"\n--- Лист '{sheet_name}' ---")
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_value = str(cell.value)
                            if search_str in cell_value:
                                found_count += 1
                                print(f"  Ячейка {cell.coordinate}: {cell_value}")
            
            print(f"\nВсего найдено: {found_count}")
            
        else:
            print(f"Неподдерживаемый формат файла: {file_ext}")
            
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")

# Использование
file_path = r'C:\glavnaya\Price.xls'  # или .xlsx
find_value_in_excel(file_path, "120/100")