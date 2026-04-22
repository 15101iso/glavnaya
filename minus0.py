import openpyxl
import re

def remove_dot_zero_force(input_file, output_file):
    """
    Принудительно удаляет .0 из значений в столбцах Артикул и ОПТ
    Работает с числами, строками, любыми форматами
    """
    
    print("=" * 70)
    print(f"🔧 ПРИНУДИТЕЛЬНОЕ УДАЛЕНИЕ .0 ИЗ ФАЙЛА: {input_file}")
    print("=" * 70)
    
    # Загружаем файл
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    
    # Находим столбцы
    article_col = None
    opt_col = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            header_lower = str(header).lower()
            if 'артикул' in header_lower or 'article' in header_lower:
                article_col = col
                print(f"✅ Найден столбец Артикул: колонка {col} ('{header}')")
            elif 'опт' in header_lower or 'opt' in header_lower or 'wholesale' in header_lower:
                opt_col = col
                print(f"✅ Найден столбец ОПТ: колонка {col} ('{header}')")
    
    if not article_col and not opt_col:
        print("❌ Столбцы не найдены!")
        return
    
    changes = 0
    
    # Обрабатываем строки
    for row in range(2, sheet.max_row + 1):
        # Обработка столбца Артикул
        if article_col:
            cell = sheet.cell(row=row, column=article_col)
            old_value = cell.value
            
            if old_value is not None:
                new_value = str(old_value).replace('.0', '')
                
                # Если значение изменилось
                if str(old_value) != new_value:
                    # Пробуем сохранить как число, если это возможно
                    try:
                        if new_value.isdigit():
                            cell.value = int(new_value)
                        else:
                            cell.value = new_value
                    except:
                        cell.value = new_value
                    
                    print(f"📌 Строка {row-1}, Артикул: '{old_value}' → '{cell.value}'")
                    changes += 1
        
        # Обработка столбца ОПТ
        if opt_col:
            cell = sheet.cell(row=row, column=opt_col)
            old_value = cell.value
            
            if old_value is not None:
                new_value = str(old_value).replace('.0', '')
                
                if str(old_value) != new_value:
                    try:
                        if new_value.isdigit():
                            cell.value = int(new_value)
                        else:
                            cell.value = new_value
                    except:
                        cell.value = new_value
                    
                    print(f"📌 Строка {row-1}, ОПТ: '{old_value}' → '{cell.value}'")
                    changes += 1
    
    # Сохраняем
    wb.save(output_file)
    
    print("\n" + "=" * 70)
    print(f"✅ Готово! Изменено ячеек: {changes}")
    print(f"✅ Результат сохранен в: {output_file}")
    print("=" * 70)

# Запуск
remove_dot_zero_force('processed_data.xlsx', 'processed_data_final.xlsx')