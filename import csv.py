import csv
from collections import defaultdict
import os

def find_duplicates_in_column(filename, column_name='артикул', output_file='duplicates_report.txt'):
    """
    Находит повторяющиеся значения в указанном столбце
    """
    
    if not os.path.exists(filename):
        print(f"❌ Файл {filename} не найден!")
        return
    
    print("=" * 70)
    print(f"ПОИСК ДУБЛИКАТОВ В СТОЛБЦЕ '{column_name}'")
    print("=" * 70)
    
    # Словарь для хранения всех значений и их позиций
    values_dict = defaultdict(list)
    headers = []
    total_rows = 0
    empty_values = 0
    
    # Пробуем разные разделители
    delimiters = [',', ';', '\t']
    file_encoding = 'utf-8'
    
    for delimiter in delimiters:
        try:
            with open(filename, 'r', encoding=file_encoding) as f:
                reader = csv.reader(f, delimiter=delimiter)
                
                # Читаем заголовки
                headers = next(reader)
                
                # Находим индекс нужного столбца
                col_index = None
                for i, h in enumerate(headers):
                    if column_name.lower() in h.lower():
                        col_index = i
                        print(f"✅ Найден столбец: '{h}' (индекс {i})")
                        break
                
                if col_index is None:
                    print(f"❌ Столбец '{column_name}' не найден!")
                    print(f"   Доступные столбцы: {', '.join(headers)}")
                    return
                
                print(f"📊 Разделитель: '{delimiter}'")
                
                # Читаем данные
                for row_num, row in enumerate(reader, 2):  # начинаем с 2 (после заголовков)
                    total_rows += 1
                    
                    if len(row) > col_index:
                        value = row[col_index].strip()
                        
                        if value:  # если не пустое
                            values_dict[value].append(row_num)
                        else:
                            empty_values += 1
                
                break  # если успешно прочитали, выходим из цикла
                
        except Exception as e:
            continue
    
    # Анализируем результаты
    duplicates = {value: rows for value, rows in values_dict.items() if len(rows) > 1}
    
    print("\n" + "=" * 70)
    print(f"РЕЗУЛЬТАТЫ ПОИСКА:")
    print("=" * 70)
    print(f"📊 Всего строк с данными: {total_rows}")
    print(f"📊 Уникальных артикулов: {len(values_dict)}")
    print(f"📊 Пустых значений: {empty_values}")
    print(f"📊 Артикулов с дубликатами: {len(duplicates)}")
    
    if duplicates:
        print("\n🔍 ДУБЛИКАТЫ НАЙДЕНЫ:")
        print("-" * 70)
        
        # Сортируем по количеству дубликатов
        sorted_duplicates = sorted(duplicates.items(), key=lambda x: len(x[1]), reverse=True)
        
        for value, rows in sorted_duplicates[:20]:  # покажем первые 20
            print(f"📌 Артикул: '{value}'")
            print(f"   Встречается {len(rows)} раз в строках: {rows}")
            print()
        
        if len(sorted_duplicates) > 20:
            print(f"... и еще {len(sorted_duplicates) - 20} артикулов")
        
        # Сохраняем отчет
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"ОТЧЕТ О ДУБЛИКАТАХ В СТОЛБЦЕ '{column_name}'\n")
            f.write("=" * 60 + "\n")
            f.write(f"Файл: {filename}\n")
            f.write(f"Всего строк: {total_rows}\n")
            f.write(f"Уникальных артикулов: {len(values_dict)}\n")
            f.write(f"Артикулов с дубликатами: {len(duplicates)}\n\n")
            
            for value, rows in sorted_duplicates:
                f.write(f"Артикул: {value}\n")
                f.write(f"  Строки: {rows}\n")
                f.write(f"  Количество: {len(rows)}\n\n")
        
        print(f"\n✅ Полный отчет сохранен в файл: {output_file}")
        
        # Создаем файл с дубликатами для просмотра
        create_duplicates_file(filename, column_name, duplicates, headers, col_index)
        
    else:
        print("\n✅ ДУБЛИКАТОВ НЕ НАЙДЕНО!")
    
    return duplicates

def create_duplicates_file(original_file, column_name, duplicates, headers, col_index):
    """Создает Excel файл со всеми строками, содержащими дубликаты"""
    
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Дубликаты"
        
        # Заголовки
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Добавляем колонку с количеством дубликатов
        ws.cell(row=1, column=len(headers)+1, value="Количество дубликатов")
        
        # Собираем все строки с дубликатами
        rows_with_duplicates = []
        
        with open(original_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            all_rows = list(reader)
        
        for value, rows in duplicates.items():
            for row_num in rows:
                if row_num <= len(all_rows):
                    row_data = all_rows[row_num-1]  # -1 потому что row_num включает заголовок
                    rows_with_duplicates.append((row_data, len(rows)))
        
        # Записываем данные
        for row_idx, (row_data, dup_count) in enumerate(rows_with_duplicates, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            ws.cell(row=row_idx, column=len(headers)+1, value=dup_count)
        
        output_file = 'duplicates_detail.xlsx'
        wb.save(output_file)
        print(f"✅ Детальный файл с дубликатами: {output_file}")
        
    except ImportError:
        print("ℹ️ Для создания Excel файла установите openpyxl: pip install openpyxl")
    except Exception as e:
        print(f"⚠️ Не удалось создать Excel файл: {e}")

# Запуск
if __name__ == "__main__":
    filename = 'allMRT.csv'
    find_duplicates_in_column(filename, 'артикул')