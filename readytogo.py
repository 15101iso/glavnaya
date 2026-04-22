import xlrd
import openpyxl
from openpyxl import Workbook
import os
import re

def parse_tire_size(value):
    """
    Разбирает строку с размером шины на составляющие
    Пример: "140/80 R18" -> (140, 80, 18)
    """
    if not value or not isinstance(value, str):
        return None, None, None
    
    # Ищем паттерн: число/число R число
    pattern = r'(\d+)[/](\d+)\s*R\s*(\d+)'
    match = re.search(pattern, value.strip())
    
    if match:
        width = int(match.group(1))      # 140
        profile = int(match.group(2))    # 80
        diameter = int(match.group(3))   # 18
        return width, profile, diameter
    
    # Альтернативный паттерн: число/число пробел число
    pattern2 = r'(\d+)[/](\d+)\s*(\d+)'
    match = re.search(pattern2, value.strip())
    
    if match:
        width = int(match.group(1))      # 140
        profile = int(match.group(2))    # 80
        diameter = int(match.group(3))   # 18
        return width, profile, diameter
    
    return None, None, None

def process_excel_file(input_file, output_file='output.xlsx', error_file='errors.txt'):
    """
    Обрабатывает Excel файл:
    - Читает данные из 7-го столбца (индекс 6)
    - Разделяет значения на ширину, профиль, диаметр
    - Создает новый файл с дополнительными столбцами
    - Записывает ошибки в отдельный файл
    """
    
    # Проверяем существование входного файла
    if not os.path.exists(input_file):
        print(f"Ошибка: Файл {input_file} не найден!")
        return False
    
    # Определяем формат входного файла по расширению
    file_ext = os.path.splitext(input_file)[1].lower()
    
    # Список для хранения ошибок
    errors = []
    
    # Данные для нового файла
    all_rows = []
    headers = None
    
    print(f"Обрабатывается файл: {input_file}")
    
    try:
        if file_ext == '.xls':
            # Обработка старого формата .xls
            import xlrd
            book = xlrd.open_workbook(input_file)
            sheet = book.sheet_by_index(0)
            
            # Получаем заголовки (первая строка)
            if sheet.nrows > 0:
                headers = [str(cell.value).strip() for cell in sheet.row(0)]
            
            # Добавляем новые заголовки
            new_headers = headers + ['Ширина', 'Профиль', 'Диаметр']
            
            # Обрабатываем данные (начиная со 2-й строки)
            for row_idx in range(1, sheet.nrows):
                row = sheet.row(row_idx)
                row_data = [str(cell.value).strip() if cell.value else '' for cell in row]
                
                # Проверяем, есть ли 7-й столбец
                if len(row_data) > 6:
                    seventh_col = row_data[6]
                    
                    # Пытаемся разобрать значение
                    width, profile, diameter = parse_tire_size(seventh_col)
                    
                    if width is not None:
                        # Успешно разобрали
                        new_row = row_data + [width, profile, diameter]
                        all_rows.append(new_row)
                    else:
                        # Ошибка разбора
                        error_msg = f"Строка {row_idx + 1}: Не удалось разобрать '{seventh_col}'"
                        errors.append(error_msg)
                        print(f"  ⚠️ {error_msg}")
                        
                        # Добавляем строку с пустыми значениями
                        new_row = row_data + ['', '', '']
                        all_rows.append(new_row)
                else:
                    # Нет 7-го столбца
                    error_msg = f"Строка {row_idx + 1}: Нет 7-го столбца"
                    errors.append(error_msg)
                    print(f"  ⚠️ {error_msg}")
                    
                    # Добавляем строку с пустыми значениями
                    new_row = row_data + ['', '', ''] if len(row_data) < 7 else row_data
                    all_rows.append(new_row)
        
        elif file_ext == '.xlsx':
            # Обработка нового формата .xlsx
            from openpyxl import load_workbook
            
            wb = load_workbook(input_file, data_only=True)
            sheet = wb.active
            
            # Получаем заголовки
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else '')
            
            # Добавляем новые заголовки
            new_headers = headers + ['Ширина', 'Профиль', 'Диаметр']
            
            # Обрабатываем данные
            for row_idx in range(2, sheet.max_row + 1):
                row_data = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(str(cell_value) if cell_value else '')
                
                # Проверяем, есть ли 7-й столбец
                if len(row_data) > 6:
                    seventh_col = row_data[6]
                    
                    # Пытаемся разобрать значение
                    width, profile, diameter = parse_tire_size(seventh_col)
                    
                    if width is not None:
                        new_row = row_data + [width, profile, diameter]
                        all_rows.append(new_row)
                    else:
                        error_msg = f"Строка {row_idx}: Не удалось разобрать '{seventh_col}'"
                        errors.append(error_msg)
                        print(f"  ⚠️ {error_msg}")
                        new_row = row_data + ['', '', '']
                        all_rows.append(new_row)
                else:
                    error_msg = f"Строка {row_idx}: Нет 7-го столбца"
                    errors.append(error_msg)
                    print(f"  ⚠️ {error_msg}")
                    new_row = row_data + ['', '', '']
                    all_rows.append(new_row)
        else:
            print(f"Неподдерживаемый формат файла: {file_ext}")
            return False
        
        # Создаем новый Excel файл с результатами
        create_output_file(output_file, new_headers, all_rows)
        
        # Записываем ошибки в файл
        write_errors_to_file(error_file, errors, input_file)
        
        # Статистика
        print(f"\n{'='*50}")
        print(f"ОБРАБОТКА ЗАВЕРШЕНА")
        print(f"{'='*50}")
        print(f"✅ Всего обработано строк: {len(all_rows)}")
        print(f"✅ Успешно разобрано: {len(all_rows) - len(errors)}")
        print(f"⚠️  Ошибок: {len(errors)}")
        print(f"📁 Результат сохранен в: {output_file}")
        print(f"📁 Ошибки сохранены в: {error_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        return False

def create_output_file(filename, headers, data):
    """Создает новый Excel файл с результатами"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Обработанные данные"
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Записываем данные
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Автоматически подгоняем ширину столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Сохраняем файл
    wb.save(filename)
    print(f"  ✅ Файл {filename} создан")

def write_errors_to_file(filename, errors, source_file):
    """Записывает ошибки в текстовый файл"""
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"ОШИБКИ ОБРАБОТКИ ФАЙЛА: {source_file}\n")
        f.write(f"{'='*50}\n")
        f.write(f"Всего ошибок: {len(errors)}\n")
        f.write(f"{'='*50}\n\n")
        
        for i, error in enumerate(errors, 1):
            f.write(f"{i}. {error}\n")
    
    print(f"  ✅ Файл {filename} создан с {len(errors)} ошибками")

def create_error_rows_file(input_file, error_file='error_rows.xlsx'):
    """
    Создает отдельный файл только со строками, содержащими ошибки
    """
    if not os.path.exists(input_file):
        return
    
    file_ext = os.path.splitext(input_file)[1].lower()
    error_rows = []
    headers = None
    
    try:
        if file_ext == '.xls':
            book = xlrd.open_workbook(input_file)
            sheet = book.sheet_by_index(0)
            
            if sheet.nrows > 0:
                headers = [str(cell.value).strip() for cell in sheet.row(0)]
            
            for row_idx in range(1, sheet.nrows):
                row = sheet.row(row_idx)
                row_data = [str(cell.value).strip() if cell.value else '' for cell in row]
                
                if len(row_data) > 6:
                    seventh_col = row_data[6]
                    width, profile, diameter = parse_tire_size(seventh_col)
                    
                    if width is None and seventh_col.strip():
                        # Это строка с ошибкой
                        error_rows.append([row_idx + 1] + row_data)
        
        elif file_ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(input_file, data_only=True)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            
            for row_idx in range(2, sheet.max_row + 1):
                row_data = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(cell_value)
                
                if len(row_data) > 6:
                    seventh_col = str(row_data[6]) if row_data[6] else ''
                    width, profile, diameter = parse_tire_size(seventh_col)
                    
                    if width is None and seventh_col.strip():
                        error_rows.append([row_idx] + row_data)
        
        if error_rows and headers:
            # Создаем файл с ошибками
            wb_error = Workbook()
            ws_error = wb_error.active
            ws_error.title = "Строки с ошибками"
            
            # Заголовки
            error_headers = ['№ строки'] + headers
            for col_idx, header in enumerate(error_headers, 1):
                ws_error.cell(row=1, column=col_idx, value=header)
            
            # Данные
            for row_idx, row_data in enumerate(error_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws_error.cell(row=row_idx, column=col_idx, value=value)
            
            wb_error.save(error_file)
            print(f"  ✅ Файл с ошибочными строками: {error_file}")
            
    except Exception as e:
        print(f"Ошибка при создании файла с ошибками: {e}")

# Основная функция
def main():
    print("="*60)
    print("ПРОГРАММА ДЛЯ РАЗДЕЛЕНИЯ ДАННЫХ В EXCEL")
    print("="*60)
    
    # Путь к входному файлу
    input_file = r'C:\glavnaya\Price.xls'  # Укажите ваш путь
    
    # Имена выходных файлов
    output_file = 'processed_data.xlsx'
    error_file = 'processing_errors.txt'
    error_rows_file = 'error_rows.xlsx'
    
    # Обработка основного файла
    success = process_excel_file(input_file, output_file, error_file)
    
    if success:
        # Создаем отдельный файл только с ошибочными строками
        create_error_rows_file(input_file, error_rows_file)
        
        print(f"\n{'='*60}")
        print("ГОТОВО!")
        print(f"{'='*60}")
        print(f"📁 Основной файл с результатами: {output_file}")
        print(f"📁 Текстовый файл с ошибками: {error_file}")
        print(f"📁 Excel файл с ошибочными строками: {error_rows_file}")
    else:
        print("\n❌ Обработка не удалась!")

if __name__ == "__main__":
    main()